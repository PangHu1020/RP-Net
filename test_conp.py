import torch
import torch.nn.functional as F
from tqdm import tqdm
import os
import pandas as pd
from pandas import ExcelWriter
import openpyxl
from openpyxl.styles import Alignment

# from model import Linearhead, CosineSimilarityLoss, CenterLoss
from model.head.mlp import MLP
from model.neck.avgpool import Avgpool
from model.backbone.resnet import ResNet
from model.network.network import load_model
from dataset.dataset import test_dataloader
from utils.metric import Metrics
from args_parser import parse_args
from utils.log import setup_logging

# Parse arguments
args = parse_args()

# Set device
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

# Define model components
model = load_model(args.name)

# Load the best model checkpoint
checkpoint_path = os.path.join(args.checkpoints, args.name, 'best.pth')
if not os.path.exists(checkpoint_path):
    raise FileNotFoundError(f"Checkpoint not found at {checkpoint_path}")

checkpoint = torch.load(checkpoint_path, map_location=device)
model.load_state_dict(checkpoint['model_state_dict']['model'])

# Set models to evaluation mode
model.eval()

# Define metric calculation
metric = Metrics(num_classes=args.num_classes)

# Evaluation loop
progress_test = tqdm(test_dataloader, desc='test')
metric.reset()

with torch.no_grad():
    for ls_images, gt_images, labels in progress_test:
        # Move data to the appropriate device
        ls_images = ls_images.to(device)
        gt_images = gt_images.to(device)
        labels = labels.to(device)

        # Forward pass
        # feature_cls = neck(E1(ls_images))
        # feature_cls = neck(E1(gt_images))
        outputs = model(ls_images)
        outputs = F.softmax(outputs, dim=1)

        # Update metrics
        metric.update(outputs, labels)

# Calculate final metrics
acc = metric.accuracy()
precision = metric.precision()
recall = metric.recall()
f1 = metric.f1_score()
specificity = metric.specificity()
g_mean = metric.g_mean()
kappa = metric.kappa()
dice = metric.dice()
auc = metric.auc_score()

variance = metric.metric_variance()
metrics_dict = {
    'Metric': ['Accuracy', 'Precision', 'Recall', 'F1 Score', 'Specificity', 'G-Mean', 'Kappa', 'Dice', 'AUC'],
    'Value': [
        f"{acc:.2f} ± {variance['accuracy_std'] * 100:.2f}",  # Accuracy with its variance multiplied by 100
        f"{precision:.2f} ± {variance['precision_std'] * 100:.2f}",
        f"{recall:.2f} ± {variance['recall_std'] * 100:.2f}",
        f"{f1:.2f} ± {variance['f1_std'] * 100:.2f}",
        f"{specificity:.2f} ± {variance['specificity_std'] * 100:.2f}",
        f"{g_mean:.2f} ± {variance['g_mean_std'] * 100:.2f}",
        f"{kappa:.2f} ± {variance['kappa_std'] * 100:.2f}",  # Kappa with its variance multiplied by 100
        f"{dice:.2f} ± {variance['dice_std'] * 100:.2f}",
        f"{auc:.2f} ± {variance['auc_std'] * 100:.2f}"   # AUC with its variance multiplied by 100
    ]
}

# 将字典转换为 pandas DataFrame
metrics_df = pd.DataFrame(metrics_dict)

# Transpose the DataFrame to make it horizontal
metrics_df = metrics_df.T  # Switch rows and columns

# 保存到 Excel 文件
excel_path = os.path.join(args.exp, args.name + '.xlsx')
metrics_df.to_excel(excel_path, index=False, header=False, engine='openpyxl')

# Load the Excel file to apply further formatting
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

# 设置单元格内容居中对齐
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# 保存调整后的 Excel 文件
wb.save(excel_path)

print("Test finish!")